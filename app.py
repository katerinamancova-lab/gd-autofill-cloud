import io
import os
import re
import json
import time
from urllib.parse import quote_plus, urlparse, parse_qs, unquote

import requests
import streamlit as st
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from dotenv import load_dotenv

load_dotenv()

try:
    import google.generativeai as genai
except Exception:
    genai = None


YELLOW = PatternFill(fill_type="solid", fgColor="FFFF00")

BRAND_DEFAULTS = {
    "союз": ("СОЮЗ", "Россия", "Россия"),
    "favorit": ("Favorit", "Россия", "Россия"),
    "honda": ("Honda", "Япония", "Япония"),
    "tohatsu": ("Tohatsu", "Япония", "Япония"),
    "suzuki": ("Suzuki", "Япония", "Япония"),
    "yamaha": ("Yamaha", "Япония", "Япония"),
    "mercury": ("Mercury", "США", "Китай"),
    "hidea": ("Hidea", "Китай", "Китай"),
    "hdx": ("HDX", "Китай", "Китай"),
    "parsun": ("Parsun", "Китай", "Китай"),
    "sea-pro": ("Sea-Pro", "Китай", "Китай"),
    "seanovo": ("Seanovo", "Китай", "Китай"),
    "reef rider": ("Reef Rider", "Китай", "Китай"),
    "bajaj": ("Bajaj", "Индия", "Индия"),
    "voge": ("VOGE", "Китай", "Китай"),
    "qjmotor": ("QJMotor", "Китай", "Китай"),
    "royal enfield": ("Royal Enfield", "Индия", "Индия"),
    "ktm": ("KTM", "Австрия", "Индия"),
    "benelli": ("Benelli", "Италия", "Китай"),
    "stels": ("Stels", "Россия", "Китай"),
    "benda": ("Benda", "Китай", "Китай"),
    "linhai": ("Linhai Yamaha", "Китай", "Китай"),
    "greencamel": ("GreenCamel", "Россия", "Китай"),
    "sprmotors": ("SPRMOTORS", "Китай", "Китай"),
    "motoland": ("Motoland", "Россия", "Китай"),
    "kayo": ("Kayo", "Китай", "Китай"),
    "regulmoto": ("Regulmoto", "Россия", "Китай"),
    "cfmoto": ("CFMOTO", "Китай", "Китай"),
    "brp": ("BRP", "Канада", "Канада"),
    "segway": ("Segway", "США", "Китай"),
}

CATEGORY_KEYS = {
    "лодочный мотор": ["дейдвуд", "вращение винта", "тип насадки", "передачи"],
    "лодка пвх": ["тип днища", "плотность материала", "диаметр борта", "внутренняя длина"],
    "квадроцикл": ["наличие псм", "лебедка", "тип привода", "защита рук", "фаркоп"],
    "гольфкар": ["пассажировместимость", "мощность (вт)", "запас хода", "педаль акселератора"],
    "мотоцикл": ["тип мотоцикла", "наличие птс", "колеса передние", "колеса задние"],
}

MOTO_ACCESSORIES = (
    "Шлем кроссовый Sharmax SH536 Red/Black;"
    "Шлем кроссовый Sharmax SH336 Blue/Black;"
    "Мотозащита Sharmax черепаха RT 8;"
    "Очки кроссовые Sharmax Premium Black;"
    "Очки кроссовые Sharmax Gray/Black;"
    "Наколенники Sharmax пластик KP 48 Красные;"
    "Наколенники Sharmax SH-32K;"
    "Мотоперчатки Sharmax GL-SH 47 White ;"
    "Мотоперчатки Sharmax GL-SH 48 Yellow;"
    "Мотоперчатки Sharmax GL-SH 49 Red\n"
)

SERVICE_DEFAULTS = {
    "Комплектация": '<ul><li style="font-size: 13pt; font-family: Acrom">Товар</li><li style="font-size: 13pt; font-family: Acrom">Документы</li></ul>',
    "Гарантия на товар": '<span style="font-family: Acrom; font-size: 13pt;">Гарантия на товар составляет 1 год</span>',
    "Скидка": 11,
    "Доступное количество": 1000,
    "Сортировка": 500,
    "Привязка к аксессуарам (новая)": MOTO_ACCESSORIES,
}


def get_secret(name: str) -> str:
    try:
        value = st.secrets.get(name, "")
    except Exception:
        value = os.getenv(name, "")
    return str(value or "").strip()


def header_map(ws):
    return {str(ws.cell(1, c).value or "").strip(): c for c in range(1, ws.max_column + 1)}


def detect_category(headers, first_names):
    joined = " ".join(str(h or "").lower() for h in headers)
    for cat, keys in CATEGORY_KEYS.items():
        if any(k in joined for k in keys):
            return cat

    names = " ".join(first_names).lower()
    if any(x in names for x in ["bf", "mfs", "mercury", "tohatsu", "hidea", "parsun", "sea-pro", "suzuki df"]):
        return "лодочный мотор"
    if any(x in names for x in ["пвх", "лодка", "нднд", "airdeck", "союз", "байкал", "сапфир"]):
        return "лодка пвх"
    if any(x in names for x in ["atv", "outlander", "segway", "linhai", "квадроцикл", "sprmotors", "blade"]):
        return "квадроцикл"
    if any(x in names for x in ["greencamel", "гольфкар", "сонора"]):
        return "гольфкар"
    return "мотоцикл"


def set_if_col(ws, hmap, row, header, value):
    if value is None:
        return 0
    col = hmap.get(header)
    if not col:
        return 0
    if any(x in header.lower() for x in ["uid", "уид", "активность", "розничная цена"]):
        return 0
    cell = ws.cell(row, col)
    if str(cell.value or "").strip() == str(value or "").strip():
        return 0
    cell.value = value
    cell.fill = YELLOW
    return 1


def apply_brand(spec, name):
    p = name.lower()
    for key, (brand, country, made) in BRAND_DEFAULTS.items():
        if key in p:
            spec.setdefault("Бренд [BRAND]", brand)
            spec.setdefault("Страна бренда [BRAND_COUNTRY]", country)
            spec.setdefault("Страна производства [MANUFACTURER]", made)
            return spec
    return spec


def range_hp_motor(hp):
    try:
        hp = float(hp)
    except Exception:
        return ""
    if hp <= 3.9:
        return "до 3.9"
    if hp <= 6.9:
        return "4 - 6.9"
    if hp <= 9.8:
        return "7 - 9.8"
    if hp <= 20:
        return "9.9 - 20"
    if hp <= 39:
        return "21 - 39"
    if hp <= 59:
        return "40 - 59"
    if hp <= 79:
        return "60 - 79"
    if hp <= 130:
        return "80 - 130"
    if hp <= 150:
        return "131 - 150"
    return "Более 151"


def hp_from_name(name):
    p = name.lower()
    for pat in [r"bf\s*([0-9]{2,3})", r"f\s*([0-9]{2,3})", r"mfs\s*([0-9]{2,3})", r"df\s*([0-9]{2,3})", r"hdef\s*([0-9]{2,3})", r"hd\s*([0-9]{2,3})", r"t\s*([0-9]{2,3})", r"([0-9]{2,3})\s*л\.?с"]:
        m = re.search(pat, p)
        if m:
            v = int(m.group(1))
            if 2 <= v <= 350:
                return v
    return None


def range_cc_moto(cc):
    try:
        cc = float(cc)
    except Exception:
        return ""
    if cc <= 199:
        return "до 199"
    if cc <= 300:
        return "200 - 300"
    if cc <= 600:
        return "301 - 600"
    if cc <= 800:
        return "601 - 800"
    return "от 801"


def range_hp_moto(hp):
    try:
        hp = float(hp)
    except Exception:
        return ""
    if hp <= 8:
        return "2 - 8"
    if hp <= 15:
        return "9 - 15"
    if hp <= 25:
        return "16 - 25"
    if hp <= 40:
        return "26 - 40"
    if hp <= 60:
        return "41 - 60"
    if hp <= 100:
        return "61 - 100"
    return "Более 100"


MOTO_KNOWN = {
    "voge ds800": (798, 94, "Тур-эндуро", "Жидкостное", "Инжектор", 2),
    "honda cb400": (399, 46, "Дорожный", "Жидкостное", "Карбюратор", 4),
    "gold wing": (1833, 126, "Туристический", "Жидкостное", "Инжектор", 6),
    "srk 921": (921, 129, "Нэйкед", "Жидкостное", "Инжектор", 4),
    "srk 800": (799, 95, "Спортивный", "Жидкостное", "Инжектор", 4),
    "srk 600": (554, 61, "Нэйкед", "Жидкостное", "Инжектор", 4),
    "srk 550": (554, 56, "Нэйкед", "Жидкостное", "Инжектор", 2),
    "srv 400": (385, 41, "Классический", "Жидкостное", "Инжектор", 2),
    "srv 550": (554, 61, "Классический", "Жидкостное", "Инжектор", 2),
    "pulsar ns400": (373, 40, "Нэйкед", "Жидкостное", "Инжектор", 1),
    "ss400": (373, 40, "Спортивный", "Жидкостное", "Инжектор", 1),
    "c4 300": (300, 26, "Дорожный", "Воздушное", "Карбюратор", 1),
    "d4 125": (125, 11, "Дорожный", "Воздушное", "Карбюратор", 1),
    "390 adventure": (373, 43, "Тур-эндуро", "Жидкостное", "Инжектор", 1),
    "390 duke": (373, 43, "Нэйкед", "Жидкостное", "Инжектор", 1),
    "himalayan 450": (452, 40, "Тур-эндуро", "Жидкостное", "Инжектор", 1),
    "guerrilla 450": (452, 40, "Нэйкед", "Жидкостное", "Инжектор", 1),
    "benelli 752": (754, 76, "Нэйкед", "Жидкостное", "Инжектор", 2),
    "hunter 350": (349, 20, "Классический", "Воздушное", "Инжектор", 1),
    "interceptor": (648, 47, "Классический", "Воздушно-масляное", "Инжектор", 2),
    "continental gt": (648, 47, "Классический", "Воздушно-масляное", "Инжектор", 2),
    "rk125": (125, 12, "Дорожный", "Воздушное", "Карбюратор", 1),
    "m502n": (500, 47, "Нэйкед", "Жидкостное", "Инжектор", 2),
    "monster plus": (125, 11, "Дорожный", "Воздушное", "Карбюратор", 1),
    "pulsar as 150": (150, 17, "Дорожный", "Воздушное", "Карбюратор", 1),
    "pulsar as 200": (199, 24, "Дорожный", "Жидкостное", "Карбюратор", 1),
    "boxer bm125": (125, 10, "Дорожный", "Воздушное", "Карбюратор", 1),
    "bajaj v 150": (150, 12, "Классический", "Воздушное", "Карбюратор", 1),
}


def rules_boat_motor(name):
    p = name.lower()
    spec = {}
    apply_brand(spec, name)
    hp = hp_from_name(name)
    if hp:
        spec["Мощность, л.с. [POWER_HP1]"] = hp
        spec["Мощность (л.с.) [POWER_HP]"] = range_hp_motor(hp)
        spec["Мощность, л.с. [POWER_HP]"] = hp
        spec["Мощность (кВт) [POWER_KW]"] = round(hp * 0.7355, 1)
    spec.setdefault("Управление [OPERATION]", "Румпельное" if any(x in p for x in ["fh", "румп", "hes"]) else "Дистанционное")
    spec.setdefault("Система запуска [STARTING_SYSTEM]", "Электростартер" if any(x in p for x in ["e", "etl", "elpt", "xrtu", "efi"]) else "Ручной стартер/электростартер")
    spec.setdefault("Дейдвуд [DEADWOOD]", "635 (XL)" if any(x in p for x in ["xrtu", "fex", "xl"]) else ("508 (L)" if any(x in p for x in ["etl", "elpt", "fel", "lrt", "fvel"]) else "381 (S)"))
    spec.setdefault("Тип насадки [NOZZLETYPE]", "Водомёт" if "jet" in p or "водом" in p else "Винт")
    spec.setdefault("Система подачи топлива [Fuel_supply_system]", "Инжектор" if "efi" in p else "Карбюратор")
    spec.setdefault("Система подъёма [LIFTING_SYSTEM]", "Гидравлическая" if any(x in p for x in ["pt", "trim", "xrtu", "elpt", "btx"]) else "Ручная")
    spec.setdefault("Количество тактов [STROKE]", "2" if "2 такт" in p or "2-такт" in p or "t 40" in p else "4")
    spec.setdefault("Охлаждение [COOLING]", "Водяное")
    spec.setdefault("Тип двигателя [TYPE_ENGINE]", "Бензиновый")
    spec.setdefault("Передачи [GEAR]", "F-N-R")
    spec.setdefault("Тип топлива [TYPE_FUEL]", "АИ92-95")
    spec.setdefault("Вращение винта [ROTATION_SCREW]", "Водомётная насадка" if spec.get("Тип насадки [NOZZLETYPE]") == "Водомёт" else "Правое")
    spec.setdefault("Гарантия [WARRANTY]", "5 лет" if any(x in p for x in ["honda", "tohatsu"]) else "3 года")
    return spec


def rules_pvc_boat(name):
    p = name.lower()
    spec = {}
    apply_brand(spec, name)
    length = re.search(r"(\d{3})", p)
    if length:
        spec["Длина, см [LENGTH_CM]"] = int(length.group(1))
    spec.setdefault("Тип лодки [TYPE_BOAT]", "Под мотор")
    spec.setdefault("Гарантия [WARRANTY]", "1 год")
    if "airdeck" in p or "нднд" in p or "air" in p or "aero" in p:
        spec["Тип днища [TYPE_BOTTOM]"] = "Надувное, низкого давления"
    elif any(x in p for x in ["фб", "фанер", "пайол"]):
        spec["Тип днища [TYPE_BOTTOM]"] = "Фанерные пайолы"
    else:
        spec["Тип днища [TYPE_BOTTOM]"] = "Надувное, низкого давления"
    spec.setdefault("Сливной клапан [DRAIN_VALVE]", "Есть")
    spec.setdefault("Надувной киль [INFLATABLE_KEEL]", "Есть")
    return spec


def rules_quad(name):
    p = name.lower()
    spec = {}
    apply_brand(spec, name)
    cc_match = re.search(r"(\d{3,4})", p)
    if cc_match:
        spec["Объём двигателя, куб [ENGINE_CAPACITY]"] = int(cc_match.group(1))
    spec.setdefault("Охлаждение [COOLING]", "Жидкостное" if any(x in p for x in ["700", "800", "1000", "efi"]) else "Воздушное")
    spec.setdefault("Тип привода [Tipprivoda]", "Полный")
    spec.setdefault("Система привода [DRIVE_SYSTEM]", "Карданный")
    spec.setdefault("Трансмиссия [TRANSMISSION]", "Вариатор")
    spec.setdefault("Наличие ПСМ [NALICHIE_PSM]", "Есть" if "псм" in p else "Нет")
    spec.setdefault("Тип двигателя [TYPE_ENGINE]", "Бензиновый")
    spec.setdefault("Система подачи топлива [Fuel_supply_system]", "Инжектор" if "efi" in p else "Карбюратор")
    spec.setdefault("Лебедка [WINCH]", "Есть")
    spec.setdefault("Классификация [CLASSIFICATION]", "Утилитарный")
    spec.setdefault("Система запуска [STARTING_SYSTEM]", "Электростартер")
    spec.setdefault("Материал рамы [FRAME_MATERIAL]", "Сталь")
    spec.setdefault("Гарантия [WARRANTY]", "1 год")
    return spec


def rules_golfcar(name):
    p = name.lower()
    spec = {}
    apply_brand(spec, name)
    spec.setdefault("Тип двигателя [TYPE_ENGINE]", "Электрический")
    spec.setdefault("Система запуска [STARTING_SYSTEM]", "Электростартер")
    spec.setdefault("Система привода", "Полный привод" if "4x4" in p else "Задний привод")
    spec.setdefault("Пассажировместимость", "4" if "2+2" in p else "2")
    spec.setdefault("Страна бренда [BRAND_COUNTRY]", "Россия")
    spec.setdefault("Страна производства [MANUFACTURER]", "Китай")
    spec.setdefault("Гарантия [WARRANTY]", "1 год")
    return spec


def rules_motorcycle(name):
    p = name.lower()
    spec = {}
    apply_brand(spec, name)
    for key, data in MOTO_KNOWN.items():
        if key in p:
            cc, hp, mtype, cooling, fuel, cylinders = data
            spec["Объём двигателя, куб [ENGINE_CAPACITY]"] = cc
            spec["Объём двигателя (по диапазонам) [ENGINE_CAPACITY1]"] = range_cc_moto(cc)
            spec["Мощность, л.с. [POWER_HP]"] = hp
            spec["Мощность (по диапазонам) [POWER_HP1]"] = range_hp_moto(hp)
            spec["Тип мотоцикла [Motorcycle_Type]"] = mtype
            spec["Охлаждение [COOLING]"] = cooling
            spec["Система подачи топлива [Fuel_supply_system]"] = fuel
            spec["Количество цилиндров [CYLINDERS]"] = cylinders
            break
    spec.setdefault("Гарантия [WARRANTY]", "1 год")
    spec.setdefault("Наличие ПТС [NALICHIE_PTS]", "Да")
    spec.setdefault("Система запуска [STARTING_SYSTEM]", "Электростартер")
    spec.setdefault("Материал рамы [FRAME_MATERIAL]", "Сталь")
    spec.setdefault("Тип топлива [TYPE_FUEL]", "АИ92-95")
    spec.setdefault("Количество тактов [STROKE]", 4)
    spec.setdefault("Тип двигателя [TYPE_ENGINE]", "Бензиновый")
    spec.setdefault("Трансмиссия [TRANSMISSION]", "Механическая")
    spec.setdefault("Карбюратор [Carburetor]", "Нет")
    return spec


def make_rules(name, category):
    if category == "лодочный мотор":
        spec = rules_boat_motor(name)
    elif category == "лодка пвх":
        spec = rules_pvc_boat(name)
    elif category == "квадроцикл":
        spec = rules_quad(name)
    elif category == "гольфкар":
        spec = rules_golfcar(name)
    else:
        spec = rules_motorcycle(name)

    for h, v in SERVICE_DEFAULTS.items():
        spec.setdefault(h, v)
    return spec


def process_excel(uploaded_file, mode):
    wb = load_workbook(uploaded_file)
    ws = wb.active
    hmap = header_map(ws)
    headers = list(hmap.keys())

    for s in ["Отчет", "Проверить", "Лог поиска"]:
        if s in wb.sheetnames:
            del wb[s]

    report = wb.create_sheet("Отчет")
    report.append(["Показатель", "Значение"])
    check = wb.create_sheet("Проверить")
    check.append(["Строка", "Товар", "Поле", "Значение", "Комментарий"])

    rows = []
    for r in range(2, ws.max_row + 1):
        name = str(ws.cell(r, 1).value or "").strip()
        if name:
            rows.append((r, name))

    if mode == "Тест: обработать первые 3 товара":
        rows = rows[:3]

    category = detect_category(headers, [n for _, n in rows[:5]])
    changed = 0
    rules_ok = 0

    progress = st.progress(0)

    for idx, (r, name) in enumerate(rows, start=1):
        row_category = detect_category(headers, [name])
        spec = make_rules(name, row_category)
        if spec:
            rules_ok += 1

        row_changed = 0
        for h, v in spec.items():
            row_changed += set_if_col(ws, hmap, r, h, v)

        changed += row_changed
        if row_changed == 0:
            check.append([r, name, "Заполнение", "", "Нечего изменить или нет колонок"])

        progress.progress(idx / max(len(rows), 1))

    for row in [
        ["Категория файла", category],
        ["Обработано товаров", len(rows)],
        ["Правила сработали", rules_ok],
        ["Изменено ячеек", changed],
        ["Комментарий", "Быстрый режим: файл возвращается всегда. Интернет/ИИ отключены, чтобы не висело."],
    ]:
        report.append(row)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


st.set_page_config(page_title="GD AutoFill Fast v23", layout="centered")
st.title("GD AutoFill Fast v23")
st.write("Быстрая версия: заполняет все строки по правилам и всегда возвращает файл. Без долгого поиска и зависаний.")

mode = st.radio(
    "Режим работы",
    ["Автоматически заполнить файл", "Тест: обработать первые 3 товара"],
    index=0,
)

uploaded = st.file_uploader("Загрузите Excel", type=["xlsx"])

if uploaded:
    st.success(f"Файл загружен: {uploaded.name}")
    try:
        _wb = load_workbook(uploaded, read_only=True)
        _ws = _wb.active
        cnt = 0
        for _r in range(2, _ws.max_row + 1):
            if str(_ws.cell(_r, 1).value or "").strip():
                cnt += 1
        st.info(f"В файле найдено товаров: {cnt}")
        uploaded.seek(0)
    except Exception:
        uploaded.seek(0)

    if st.button("Заполнить и скачать"):
        with st.spinner("Заполняю файл..."):
            try:
                result = process_excel(uploaded, mode)
                st.success("Готово")
                st.download_button(
                    "Скачать заполненный Excel",
                    data=result,
                    file_name=uploaded.name.replace(".xlsx", "_FAST_v23.xlsx"),
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            except Exception as e:
                st.error(f"Ошибка: {e}")
