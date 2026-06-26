"""Workbook inspection, safe filling and report sheet generation."""

from __future__ import annotations

import io
import re
from dataclasses import dataclass
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from config import PROTECTED_COLUMN_MARKERS, REPORT_SHEETS

YELLOW_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="17365D")
HEADER_FONT = Font(color="FFFFFF", bold=True)


@dataclass
class WorkbookLayout:
    sheet: Worksheet
    header_row: int
    headers: list[str]
    name_column: int
    data_rows: list[int]
    excluded_columns: set[int]


def _text(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


EMPTY_MARKERS = {
    "",
    "бренд",
    "заполните в админке",
    "заполните вручную",
    "заполните вручную в админке",
    "нет данных",
    "не заполнено",
    "n/a",
    "-",
}


def is_effectively_empty(value: object) -> bool:
    if value is None:
        return True
    return _text(value).lower() in EMPTY_MARKERS


def _name_score(header: str) -> int:
    normalized = header.lower()
    exact = {
        "наименование": 100,
        "название": 95,
        "наименование товара": 110,
        "название товара": 105,
        "name": 90,
        "product name": 105,
        "модель": 80,
    }
    return exact.get(normalized, 60 if "наимен" in normalized or "назван" in normalized else 0)


def _is_product_placeholder(value: object) -> bool:
    normalized = _text(value).lower()
    if not normalized:
        return False
    return (
        "бренд модель" in normalized
        or normalized in {"наименование товара", "название товара", "модель товара"}
    )


def inspect_workbook(file_bytes: bytes) -> tuple[Any, WorkbookLayout]:
    workbook = load_workbook(io.BytesIO(file_bytes))
    candidates: list[WorkbookLayout] = []
    empty_template_found = False
    for sheet in workbook.worksheets:
        if sheet.title in REPORT_SHEETS:
            continue
        for row in range(1, min(sheet.max_row, 30) + 1):
            headers = [_text(sheet.cell(row, col).value) for col in range(1, sheet.max_column + 1)]
            if sum(bool(header) for header in headers) < 2:
                continue
            scored = [(index + 1, _name_score(header)) for index, header in enumerate(headers)]
            name_column, score = max(scored, key=lambda item: item[1])
            if score <= 0:
                continue
            data_rows = [
                index
                for index in range(row + 1, sheet.max_row + 1)
                if _text(sheet.cell(index, name_column).value)
                and not _is_product_placeholder(sheet.cell(index, name_column).value)
            ]
            shifted_template = (
                name_column > 1
                and "заполнить в админке" in headers[0].lower()
            )
            excluded_columns = set(range(1, name_column)) if shifted_template else set()
            if shifted_template and not data_rows:
                empty_template_found = True
            if data_rows:
                candidates.append(
                    WorkbookLayout(
                        sheet,
                        row,
                        headers,
                        name_column,
                        data_rows,
                        excluded_columns,
                    )
                )
    if not candidates:
        if empty_template_found:
            raise ValueError(
                "Это пустой шаблон нового товара. Замените в ячейке B2 текст-пример "
                "«БРЕНД Модель» на точное название товара и загрузите файл снова."
            )
        raise ValueError(
            "Не найден лист с колонкой названия товара. Ожидается заголовок вроде "
            "«Наименование товара», «Название» или «Модель»."
        )
    layout = max(candidates, key=lambda item: len(item.data_rows))
    return workbook, layout


def is_protected(header: str) -> bool:
    normalized = header.strip().lower()
    return any(marker == normalized or marker in normalized for marker in PROTECTED_COLUMN_MARKERS)


NON_RESEARCH_COLUMN_MARKERS = {
    "скидка",
    "доступное количество",
    "нет в продаже",
    "сортировка",
    "привязка к аксессуарам",
    "гарантия на товар",
    "по диапазонам",
}


def is_non_research_column(header: str) -> bool:
    """Admin/catalog fields that should not be searched on external websites."""
    normalized = _text(header).lower()
    return any(marker in normalized for marker in NON_RESEARCH_COLUMN_MARKERS)


def writable_columns(layout: WorkbookLayout, row: int) -> list[str]:
    output = []
    for index, header in enumerate(layout.headers, start=1):
        if (
            not header
            or index == layout.name_column
            or index in layout.excluded_columns
            or is_protected(header)
            or is_non_research_column(header)
        ):
            continue
        if is_effectively_empty(layout.sheet.cell(row, index).value):
            output.append(header)
    return output


def write_values(layout: WorkbookLayout, row: int, values: dict[str, dict[str, str]]) -> int:
    header_to_columns: dict[str, list[int]] = {}
    for index, header in enumerate(layout.headers, start=1):
        if header:
            header_to_columns.setdefault(header, []).append(index)
    changed = 0
    for header, item in values.items():
        columns = header_to_columns.get(header, [])
        if not columns or is_protected(header):
            continue
        for column in columns:
            cell = layout.sheet.cell(row, column)
            if not is_effectively_empty(cell.value):
                continue
            cell.value = item["value"]
            cell.fill = YELLOW_FILL
            changed += 1
    return changed


def _replace_sheet(workbook: Any, name: str) -> Worksheet:
    if name in workbook.sheetnames:
        del workbook[name]
    return workbook.create_sheet(name)


def _write_table(sheet: Worksheet, headers: list[str], rows: list[list[Any]]) -> None:
    sheet.append(headers)
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True)
    for row in rows:
        sheet.append(row)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in sheet.columns:
        letter = column[0].column_letter
        width = min(70, max(12, max(len(_text(cell.value)) for cell in column) + 2))
        sheet.column_dimensions[letter].width = width
        for cell in column:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def add_report_sheets(
    workbook: Any,
    report_rows: list[list[Any]],
    review_rows: list[list[Any]],
    source_rows: list[list[Any]],
) -> None:
    report = _replace_sheet(workbook, "Отчёт")
    _write_table(
        report,
        [
            "Строка товара",
            "Название товара",
            "Категория",
            "Уверенность категории",
            "Полей заполнено",
            "URL найдено",
            "Firecrawl успешно прочитано",
            "Символов текста собрано",
            "Полей вернул Gemini",
            "Реально записано в Excel",
            "Сайтов пропущено",
            "Полей на проверку",
        ],
        report_rows,
    )
    review = _replace_sheet(workbook, "Проверить")
    _write_table(
        review,
        ["Строка товара", "Название товара", "Поле", "Причина", "Спорные значения"],
        review_rows,
    )
    sources = _replace_sheet(workbook, "Источники")
    _write_table(
        sources,
        [
            "Строка товара",
            "Название товара",
            "URL источника",
            "Провайдер поиска",
            "Статус",
            "Длина текста",
            "Полей извлечено Gemini",
            "Использован",
            "Ошибка",
        ],
        source_rows,
    )


def workbook_bytes(workbook: Any) -> bytes:
    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()
